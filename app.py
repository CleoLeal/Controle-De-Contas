#bibliotecas
import customtkinter as ctk 
from tkinter import *
from tkinter import messagebox
import openpyxl, xlrd
import pathlib
from openpyxl import Workbook


#setando a aparencia padrão do sistema
ctk.set_appearance_mode("System")
ctk.set_default_color_theme("dark-blue")

class App(ctk.CTk):
    #todas a funções que estiverem aqui dentro vão rodar no init
    def __init__(self):
        #significa que init está superior
        super().__init__()
        self.layout_config()
        self.apparence_theme()
        self.system()

    #configurando o layout
    def layout_config(self):
        self.title("Contas vovó")
        self.geometry("700x500")

    #para a mudança de tema do app
    def apparence_theme(self):
        self.lb_apm = ctk.CTkLabel(self, text="Tema", bg_color="transparent", text_color=['#000', '#fff']).place(x=50, y=430)
        self.opt_apm = ctk.CTkOptionMenu(self, values=["Sistema", "Dark", "Light"], command=self.change_apm).place(x=50, y=460)

    def change_apm(self, new_apparence_modes):
        ctk.set_appearance_mode(new_apparence_modes)

    #todo o sistema
    def system(self):
        #título
        ctk.CTkLabel(self, text="Sitema de controle de contas", font=("Times new roman", 24), text_color=["#000","#fff"]).place(x=200, y=5)
        ctk.CTkLabel(self, text="Por favor, preencha todos os campos do formulário!", font=("Times new roman", 16), text_color=["#000","#fff"]).place(x=50, y=50)

        #labels
        lb_conta = ctk.CTkLabel(self, text="Conta: *", font=("Times new roman", 16), text_color=["#000","#fff"])
        lb_data = ctk.CTkLabel(self, text="Data: *", font=("Times new roman", 16), text_color=["#000","#fff"])
        lb_valor = ctk.CTkLabel(self, text="Valor: *", font=("Times new roman", 16), text_color=["#000","#fff"])
        lb_observacao = ctk.CTkLabel(self, text="Observação:", font=("Times new roman", 16), text_color=["#000","#fff"])


        #variaveis entry
        data_value = StringVar()
        valor_value =StringVar()

        #nome do excel
        ficheiro = pathlib.Path("Conta.xlsx")
        
        #se o excel existir
        if ficheiro.exists():
            #ele somente abre ele
            ficheiro = openpyxl.load_workbook("Conta.xlsx")
            folha = ficheiro.active
            folha['A1'] = "Conta"
            folha['B1'] = "Valor"
            folha['C1'] = "Data"
            folha['D1'] = "Observação"
            #e salva os dados
            ficheiro.save("Conta.xlsx")
        #senão
        else:
            #cria um excel
            ficheiro = Workbook()
            #informa que ele está ativo e dá nomes para algumas células
            folha = ficheiro.active
            folha['A1'] = "Conta"
            folha['B1'] = "Valor"
            folha['C1'] = "Data"
            folha['D1'] = "Observação"
            #salva o excel
            ficheiro.save("Conta.xlsx")

        #função do botão de enviar os dados para o excel
        def Submit():

            #pegando os dados dos entrys
            conta = cb_conta.get()
            valor = valor_value.get()
            data = data_value.get()
            obs = txt_obsevacoes.get(0.0, END)

            #se algum desses abaixo estiver vazio. Ele informa o erro e não salva os dados no excel
            if conta == "" or valor == "" or data == "":
                #mensagem
                messagebox.showerror("Sistema","ERRO!\nPor favor, preencha todos os campos com *")
            #senão
            else:
                #ele adiciona nas células os valores
                folha.cell(column=1, row=folha.max_row+1, value=conta)
                folha.cell(column=2, row=folha.max_row, value=valor)
                folha.cell(column=3, row=folha.max_row, value=data)
                folha.cell(column=4, row=folha.max_row, value=obs)
                #salva o excel
                ficheiro.save('Conta.xlsx')
                #e mostra uma mensagem que foi adicionado
                messagebox.showinfo("Sistema", "Conta paga adicionada com sucesso!")

                #depois disso, o sistema limpa as caixas de textos
                Clear()
        #função do botão de limpar as caixas de texto
        def Clear():
            cb_conta.set("")
            valor_value.set("")
            data_value.set("")
            txt_obsevacoes.delete(0.0, END)

        #entry
        entry_data = ctk.CTkEntry(self, width=210, textvariable=data_value, font=("Calibri bold", 16), fg_color="transparent", placeholder_text="Insira a data que você pagou")
        entry_valor = ctk.CTkEntry(self, textvariable=valor_value, width=180, font=("Calibri bold", 16), fg_color="transparent", placeholder_text="Insira o valor da conta")

        
        #comboBox
        cb_conta = ctk.CTkComboBox(self, width=200,values=["Alguel + Condomínio", "Gás", "Água", "Celular Vô", "Claro", "Água Casa 2", "Luz Casa 2", "IPTU Casa 2"], font=("Calibri bold", 16), fg_color="#fff", text_color="#000")
        cb_conta.set("Selecione...")

        #textBox
        txt_obsevacoes = ctk.CTkTextbox(self, width=250, height=100, font=("Arial bold", 20), border_color="#aaa", border_width=2, fg_color="transparent")

        #posicionando os elementos no frame
        lb_conta.place(x=70, y=100)
        cb_conta.place(x=70, y=125)

        lb_valor.place(x=70, y=180)
        entry_valor.place(x=70, y=205)

        lb_data.place(x= 350, y=100)
        entry_data.place(x= 350, y=125)

        lb_observacao.place(x=350, y= 180)
        txt_obsevacoes.place(x=350, y= 205)

        
        #botão de salvar os dados
        ctk.CTkButton(self, width=50, text="Salvar".upper(), font=("Calibri bold", 16), command=Submit, fg_color="#151", hover_color="#131", text_color="#000").place(x=300, y=420)
        #botão de limpar os dados
        ctk.CTkButton(self, width=50, text="Limpar".upper(), font=("Calibri bold", 16), command=Clear, fg_color="#f01", hover_color="#f21", text_color="#000").place(x=500, y=420)


if __name__=="__main__":
    app = App()
    app.mainloop()